# ============================================================
# 1_ENTRENAR_MODELO.py
# ============================================================
#
#  Lee un Excel con comentarios etiquetados y entrena un
#  modelo de clasificación NLP para categorías TELCO.
#
#  RESULTADO:
#    • CAT_CX_MODEL.pkl         → modelo listo para usar
#    • /GRAFICAS                → visualizaciones del modelo
#    • NO_CATEGORIZADOS.xlsx    → dos pestañas:
#        - Categorizados        → predicciones con alta confianza
#        - No_Categorizados     → baja confianza, para revisión/reentrenamiento
#
#  DEPENDENCIAS:
#    pip install pandas openpyxl scikit-learn joblib matplotlib seaborn
#    python -m spacy download es_core_news_sm   (opcional)
#
#  USO:
#    python 1_entrenar_modelo.py
# ============================================================

import re
import sys
import logging
import warnings
import traceback
from pathlib import Path
from datetime import datetime

import numpy as np
import pandas as pd
import joblib
warnings.filterwarnings("ignore")

from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.decomposition import TruncatedSVD
from sklearn.linear_model import LogisticRegression
from sklearn.model_selection import train_test_split, cross_val_score, StratifiedKFold
from sklearn.preprocessing import LabelEncoder
from sklearn.metrics import (
    accuracy_score, f1_score, classification_report,
    confusion_matrix, ConfusionMatrixDisplay
)
from sklearn.pipeline import Pipeline

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import matplotlib.gridspec as gridspec
from matplotlib.colors import LinearSegmentedColormap
import seaborn as sns

from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ── spaCy opcional ───────────────────────────────────────────
try:
    import spacy
    try:
        nlp = spacy.load("es_core_news_sm", disable=["ner", "parser"])
        HAS_SPACY = True
    except OSError:
        HAS_SPACY = False
        nlp = None
except ImportError:
    HAS_SPACY = False
    nlp = None

# ── SMOTE opcional ────────────────────────────────────────────
try:
    from imblearn.over_sampling import SMOTE
    HAS_SMOTE = True
except ImportError:
    HAS_SMOTE = False


# ============================================================
# ⚙️  CONFIGURACIÓN — AJUSTA ESTAS RUTAS
# ============================================================

CFG = {
    # ── Rutas ────────────────────────────────────────────────
    "archivo_entrenamiento": r"C:\Users\1872488\Documents\Model TRN\Data TR\Data de Entrenamiento.xlsx",
    "col_texto":             "Frase_clave",
    "col_categoria":         "Categoría",
    "ruta_salida_pkl":       r"C:\Users\1872488\Documents\Model TRN\RUN MODEL\CAT_CX_MODELV2.pkl",
    "ruta_graficas":         r"C:\Users\1872488\Documents\Model TRN\GRAFICAS",
    "ruta_no_categ":         r"C:\Users\1872488\Documents\Model TRN\RUN MODEL\NO_CATEGORIZADOS.xlsx",

    # ── Entrenamiento ─────────────────────────────────────────
    "test_size":          0.20,
    "random_state":       42,
    "min_muestras_clase": 10,
    "usar_lematizacion":  True,
    "usar_smote":         True,
    "tfidf_max_features": 10000,
    "tfidf_ngram_range":  (1, 2),
    "tfidf_min_df":       2,
    "tfidf_max_df":       0.95,
    "svd_components":     100,

    # ── Umbrales de confianza para "no categorizado" ─────────
    # prob_max < umbral_prob_min  → LOW_PROB
    # margen top1-top2 < umbral_margen_top → LOW_MARGIN
    # cualquiera de los dos → va a pestaña No_Categorizados
    "umbral_prob_min":    0.30,
    "umbral_margen_top":  0.15,
}


# ============================================================
# ESTILO VISUAL GLOBAL
# ============================================================

PALETTE = {
    "bg":       "#0F1117",
    "surface":  "#1A1D27",
    "accent1":  "#00D4FF",
    "accent2":  "#7B61FF",
    "accent3":  "#FF6B6B",
    "accent4":  "#2ECC71",
    "text":     "#E8EAF0",
    "muted":    "#6C7A8D",
}

CMAP_CONF = LinearSegmentedColormap.from_list(
    "telco", ["#0F1117", "#7B61FF", "#00D4FF"], N=256
)

def aplicar_estilo():
    plt.rcParams.update({
        "figure.facecolor":  PALETTE["bg"],
        "axes.facecolor":    PALETTE["surface"],
        "axes.edgecolor":    PALETTE["muted"],
        "axes.labelcolor":   PALETTE["text"],
        "axes.titlecolor":   PALETTE["text"],
        "xtick.color":       PALETTE["muted"],
        "ytick.color":       PALETTE["muted"],
        "text.color":        PALETTE["text"],
        "grid.color":        "#252A38",
        "grid.linestyle":    "--",
        "grid.alpha":        0.6,
        "legend.facecolor":  PALETTE["surface"],
        "legend.edgecolor":  PALETTE["muted"],
        "font.family":       "monospace",
        "font.size":         10,
    })

def guardar_fig(fig, nombre: str, ruta_graficas: Path, log):
    ruta = ruta_graficas / nombre
    fig.savefig(ruta, dpi=150, bbox_inches="tight",
                facecolor=PALETTE["bg"])
    plt.close(fig)
    log.info(f"   📊 Guardado: {ruta.name}")


# ============================================================
# GRÁFICA 1: Distribución de categorías
# ============================================================

def grafica_distribucion(df, col_categoria, ruta_graficas, log):
    aplicar_estilo()
    conteo = df[col_categoria].value_counts().sort_values()
    n = len(conteo)

    colores = plt.cm.cool(np.linspace(0.2, 1.0, n))
    fig, ax = plt.subplots(figsize=(12, max(5, n * 0.55)))
    fig.patch.set_facecolor(PALETTE["bg"])

    bars = ax.barh(conteo.index, conteo.values, color=colores,
                   edgecolor="none", height=0.65)

    for bar, val in zip(bars, conteo.values):
        ax.text(val + max(conteo.values) * 0.01, bar.get_y() + bar.get_height() / 2,
                f"{val:,}", va="center", ha="left",
                color=PALETTE["text"], fontsize=9, fontweight="bold")

    ax.set_xlabel("Número de muestras", labelpad=10)
    ax.set_title("Distribución de Categorías en datos de entrenamiento",
                 fontsize=14, fontweight="bold", pad=18, color=PALETTE["accent1"])
    ax.set_xlim(0, max(conteo.values) * 1.18)
    ax.grid(axis="x")
    ax.spines[["top", "right", "left"]].set_visible(False)

    total = conteo.sum()
    fig.text(0.98, 0.02, f"Total: {total:,} muestras · {n} categorías",
             ha="right", va="bottom", color=PALETTE["muted"], fontsize=8)

    guardar_fig(fig, "01_distribucion_categorias.png", ruta_graficas, log)


# ============================================================
# GRÁFICA 2: Matriz de confusión
# ============================================================

def grafica_confusion(y_test, y_pred, clases, ruta_graficas, log):
    aplicar_estilo()
    cm = confusion_matrix(y_test, y_pred)
    n  = len(clases)

    fig, ax = plt.subplots(figsize=(max(8, n * 0.9), max(7, n * 0.8)))
    fig.patch.set_facecolor(PALETTE["bg"])

    im = ax.imshow(cm, cmap=CMAP_CONF, aspect="auto")

    thresh = cm.max() / 2.0
    for i in range(n):
        for j in range(n):
            v = cm[i, j]
            color = "white" if v < thresh else PALETTE["bg"]
            if v > 0:
                ax.text(j, i, str(v), ha="center", va="center",
                        color=color, fontsize=8, fontweight="bold")

    ax.set_xticks(range(n))
    ax.set_yticks(range(n))
    ax.set_xticklabels(clases, rotation=45, ha="right", fontsize=8)
    ax.set_yticklabels(clases, fontsize=8)
    ax.set_xlabel("Predicho", labelpad=10, fontweight="bold")
    ax.set_ylabel("Real", labelpad=10, fontweight="bold")
    ax.set_title("Matriz de Confusión — Conjunto de Prueba",
                 fontsize=13, fontweight="bold", pad=18, color=PALETTE["accent1"])

    cbar = fig.colorbar(im, ax=ax, shrink=0.8)
    cbar.ax.tick_params(colors=PALETTE["muted"])

    guardar_fig(fig, "02_matriz_confusion.png", ruta_graficas, log)


# ============================================================
# GRÁFICA 3: F1 por categoría
# ============================================================

def grafica_f1_por_categoria(y_test, y_pred, clases, ruta_graficas, log):
    aplicar_estilo()
    report = classification_report(y_test, y_pred, target_names=clases,
                                   output_dict=True, zero_division=0)

    cats   = [c for c in clases if c in report]
    f1s    = [report[c]["f1-score"]  for c in cats]
    precs  = [report[c]["precision"] for c in cats]
    recs   = [report[c]["recall"]    for c in cats]
    sup    = [report[c]["support"]   for c in cats]

    orden = sorted(range(len(cats)), key=lambda i: f1s[i], reverse=True)
    cats  = [cats[i] for i in orden]
    f1s   = [f1s[i]  for i in orden]
    precs = [precs[i] for i in orden]
    recs  = [recs[i]  for i in orden]
    sup   = [sup[i]   for i in orden]

    n = len(cats)
    x = np.arange(n)
    w = 0.26

    fig, ax = plt.subplots(figsize=(max(10, n * 0.75), 6))
    fig.patch.set_facecolor(PALETTE["bg"])

    b1 = ax.bar(x - w,   precs, w, label="Precision", color=PALETTE["accent2"], alpha=0.85)
    b2 = ax.bar(x,       recs,  w, label="Recall",    color=PALETTE["accent4"], alpha=0.85)
    b3 = ax.bar(x + w,   f1s,   w, label="F1-Score",  color=PALETTE["accent1"], alpha=0.95)

    ax.axhline(0.80, color=PALETTE["accent3"], lw=1.2, ls="--", alpha=0.7, label="Meta 0.80")

    for bar, v in zip(b3, f1s):
        ax.text(bar.get_x() + bar.get_width() / 2, v + 0.015,
                f"{v:.2f}", ha="center", va="bottom",
                color=PALETTE["accent1"], fontsize=7.5, fontweight="bold")

    ax.set_xticks(x)
    ax.set_xticklabels(cats, rotation=40, ha="right", fontsize=8)
    ax.set_ylim(0, 1.15)
    ax.set_ylabel("Score", labelpad=8)
    ax.set_title("Precision / Recall / F1 por Categoría",
                 fontsize=13, fontweight="bold", pad=16, color=PALETTE["accent1"])
    ax.legend(loc="upper right", fontsize=9)
    ax.grid(axis="y")
    ax.spines[["top", "right"]].set_visible(False)

    for i, (xi, s) in enumerate(zip(x, sup)):
        ax.text(xi, -0.09, f"n={s}", ha="center", va="top",
                color=PALETTE["muted"], fontsize=7,
                transform=ax.get_xaxis_transform())

    guardar_fig(fig, "03_f1_por_categoria.png", ruta_graficas, log)


# ============================================================
# GRÁFICA 4: Curva de aprendizaje
# ============================================================

def grafica_curva_aprendizaje(modelo, X_train, y_train, ruta_graficas, log):
    from sklearn.model_selection import learning_curve
    aplicar_estilo()

    log.info("   Calculando curva de aprendizaje...")
    train_sizes, train_scores, val_scores = learning_curve(
        modelo, X_train, y_train,
        cv=5, scoring="f1_macro",
        train_sizes=np.linspace(0.1, 1.0, 8),
        n_jobs=-1
    )

    tr_mean  = train_scores.mean(axis=1)
    tr_std   = train_scores.std(axis=1)
    val_mean = val_scores.mean(axis=1)
    val_std  = val_scores.std(axis=1)

    fig, ax = plt.subplots(figsize=(9, 5))
    fig.patch.set_facecolor(PALETTE["bg"])

    ax.plot(train_sizes, tr_mean,  "o-", color=PALETTE["accent2"],
            lw=2, label="Entrenamiento", ms=6)
    ax.fill_between(train_sizes, tr_mean - tr_std, tr_mean + tr_std,
                    alpha=0.15, color=PALETTE["accent2"])

    ax.plot(train_sizes, val_mean, "o-", color=PALETTE["accent1"],
            lw=2, label="Validación (CV)", ms=6)
    ax.fill_between(train_sizes, val_mean - val_std, val_mean + val_std,
                    alpha=0.15, color=PALETTE["accent1"])

    ax.axhline(0.80, color=PALETTE["accent3"], lw=1.2, ls="--",
               alpha=0.7, label="Meta 0.80")

    ax.set_xlabel("Muestras de entrenamiento", labelpad=8)
    ax.set_ylabel("F1-Macro", labelpad=8)
    ax.set_title("Curva de Aprendizaje",
                 fontsize=13, fontweight="bold", pad=16, color=PALETTE["accent1"])
    ax.set_ylim(0, 1.05)
    ax.legend(fontsize=9)
    ax.grid(True)
    ax.spines[["top", "right"]].set_visible(False)

    guardar_fig(fig, "04_curva_aprendizaje.png", ruta_graficas, log)


# ============================================================
# GRÁFICA 5: Dashboard de métricas
# ============================================================

def grafica_dashboard(metadata, cv_scores, ruta_graficas, log):
    aplicar_estilo()
    fig = plt.figure(figsize=(14, 7))
    fig.patch.set_facecolor(PALETTE["bg"])

    gs = gridspec.GridSpec(2, 4, figure=fig, hspace=0.55, wspace=0.35)

    kpis = [
        ("Accuracy",    metadata["accuracy"],    PALETTE["accent1"]),
        ("F1-Macro",    metadata["f1_macro"],    PALETTE["accent2"]),
        ("F1-Weighted", metadata["f1_weighted"], PALETTE["accent4"]),
        ("CV F1-Macro", metadata["cv_f1_macro"], PALETTE["accent3"]),
    ]

    for idx, (label, val, color) in enumerate(kpis):
        ax = fig.add_subplot(gs[0, idx])
        ax.set_facecolor(PALETTE["surface"])

        ax.add_patch(mpatches.FancyBboxPatch(
            (0.05, 0.05), 0.90, 0.90,
            boxstyle="round,pad=0.05",
            facecolor=PALETTE["surface"],
            edgecolor=color, lw=1.5,
            transform=ax.transAxes, clip_on=False
        ))

        ax.text(0.5, 0.72, label, ha="center", va="center",
                transform=ax.transAxes, fontsize=9,
                color=PALETTE["muted"], fontweight="bold")
        ax.text(0.5, 0.38, f"{val:.4f}", ha="center", va="center",
                transform=ax.transAxes, fontsize=22,
                color=color, fontweight="bold")

        ax.add_patch(mpatches.FancyBboxPatch(
            (0.1, 0.12), 0.80 * val, 0.08,
            boxstyle="square,pad=0",
            facecolor=color, alpha=0.6,
            transform=ax.transAxes, clip_on=True
        ))
        ax.add_patch(mpatches.FancyBboxPatch(
            (0.1, 0.12), 0.80, 0.08,
            boxstyle="square,pad=0",
            facecolor="none", edgecolor=PALETTE["muted"],
            lw=0.8, transform=ax.transAxes, clip_on=True
        ))

        ax.set_xlim(0, 1); ax.set_ylim(0, 1)
        ax.axis("off")

    ax2 = fig.add_subplot(gs[1, :2])
    bplot = ax2.boxplot(
        cv_scores, vert=False, patch_artist=True,
        boxprops=dict(facecolor=PALETTE["accent2"], alpha=0.6),
        medianprops=dict(color=PALETTE["accent1"], lw=2.5),
        whiskerprops=dict(color=PALETTE["muted"]),
        capprops=dict(color=PALETTE["muted"]),
        flierprops=dict(markerfacecolor=PALETTE["accent3"],
                        marker="o", markersize=5)
    )
    ax2.scatter(cv_scores, [1] * len(cv_scores),
                color=PALETTE["accent1"], zorder=5, s=40, alpha=0.8)
    ax2.set_xlabel("F1-Macro")
    ax2.set_title("Distribución CV (5-Fold)", fontsize=10,
                  fontweight="bold", color=PALETTE["accent1"])
    ax2.set_yticks([])
    ax2.set_xlim(max(0, cv_scores.min() - 0.05), min(1, cv_scores.max() + 0.05))
    ax2.grid(axis="x")
    ax2.spines[["top", "right", "left"]].set_visible(False)

    ax3 = fig.add_subplot(gs[1, 2:])
    ax3.axis("off")

    info = [
        ["Modelo",        metadata.get("nombre", "LogReg")],
        ["Categorías",    str(len(metadata.get("categorias", [])))],
        ["Vocabulario",   f"{metadata.get('vocabulario', 0):,}"],
        ["Muestras",      f"{metadata.get('n_entrenamiento', 0):,}"],
        ["SVD comps.",    str(metadata.get("svd_componentes", 0))],
        ["Fecha",         metadata.get("fecha", "")[:10]],
    ]

    tbl = ax3.table(
        cellText=info,
        colLabels=["Parámetro", "Valor"],
        loc="center", cellLoc="left"
    )
    tbl.auto_set_font_size(False)
    tbl.set_fontsize(9)
    tbl.scale(1.1, 1.6)

    for (r, c), cell in tbl.get_celld().items():
        cell.set_facecolor(PALETTE["surface"] if r > 0 else PALETTE["bg"])
        cell.set_edgecolor(PALETTE["muted"])
        cell.set_text_props(
            color=PALETTE["accent1"] if c == 0 and r > 0 else PALETTE["text"]
        )

    ax3.set_title("Resumen del Modelo", fontsize=10,
                  fontweight="bold", color=PALETTE["accent1"])

    fig.suptitle("📊 Dashboard de Evaluación — Modelo NLP TELCO",
                 fontsize=15, fontweight="bold",
                 color=PALETTE["text"], y=1.01)

    guardar_fig(fig, "05_dashboard_metricas.png", ruta_graficas, log)


# ============================================================
# GRÁFICA 6: Top palabras por categoría
# ============================================================

def grafica_top_palabras(vectorizer, modelo, le, ruta_graficas, log, top_n=8):
    aplicar_estilo()
    clases = le.classes_
    n_cat  = len(clases)
    cols   = min(4, n_cat)
    rows   = (n_cat + cols - 1) // cols

    fig, axes = plt.subplots(rows, cols,
                             figsize=(cols * 3.8, rows * 3.2))
    fig.patch.set_facecolor(PALETTE["bg"])
    axes_flat = np.array(axes).flatten() if n_cat > 1 else [axes]

    vocab     = np.array(vectorizer.get_feature_names_out())
    palette_c = plt.cm.cool(np.linspace(0.3, 1.0, top_n))

    for idx, (ax, clase) in enumerate(zip(axes_flat, clases)):
        coefs = modelo.coef_[idx]
        top_i = np.argsort(coefs)[-top_n:][::-1]
        words = vocab[top_i]
        vals  = coefs[top_i]

        bars = ax.barh(range(top_n), vals[::-1],
                       color=palette_c, edgecolor="none", height=0.65)
        ax.set_yticks(range(top_n))
        ax.set_yticklabels(words[::-1], fontsize=7.5)
        ax.set_title(clase, fontsize=9, fontweight="bold",
                     color=PALETTE["accent1"], pad=6)
        ax.set_xlabel("Coef.", fontsize=7, labelpad=4)
        ax.spines[["top", "right"]].set_visible(False)
        ax.grid(axis="x", alpha=0.4)
        ax.tick_params(axis="x", labelsize=7)

    for ax in axes_flat[n_cat:]:
        ax.set_visible(False)

    fig.suptitle("Top Palabras Discriminantes por Categoría (coeficientes LR)",
                 fontsize=12, fontweight="bold",
                 color=PALETTE["text"], y=1.01)

    plt.tight_layout()
    guardar_fig(fig, "06_top_palabras_categoria.png", ruta_graficas, log)


# ============================================================
# LOGGING
# ============================================================

def setup_logger() -> logging.Logger:
    logging.basicConfig(
        level=logging.INFO,
        format="[%(asctime)s] %(levelname)s — %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
        handlers=[logging.StreamHandler(sys.stdout)]
    )
    return logging.getLogger("entrenar_modelo")


# ============================================================
# PREPROCESAMIENTO DE TEXTO
# ============================================================

_STOPWORDS_ES = {
    "de","la","que","el","en","y","a","los","del","se","las","un","por",
    "con","una","su","para","es","al","lo","como","más","pero","sus","le",
    "ya","o","fue","este","ha","si","porque","esta","son","entre","está",
    "cuando","muy","sin","sobre","ser","tiene","también","me","hasta","hay",
    "donde","han","quien","están",
}

def limpiar_texto(texto: str, usar_lematizacion: bool = True) -> str:
    if pd.isna(texto) or str(texto).strip() == "":
        return ""
    texto = str(texto).lower()
    texto = re.sub(r"http\S+|www\S+", " ", texto)
    texto = re.sub(r"[^\w\sáéíóúñü]", " ", texto)
    texto = re.sub(r"\d+", " ", texto)
    texto = re.sub(r"\s+", " ", texto).strip()

    if HAS_SPACY and usar_lematizacion and nlp is not None:
        doc = nlp(texto)
        tokens = [t.lemma_ for t in doc
                  if t.lemma_ not in _STOPWORDS_ES and len(t.lemma_) > 2]
    else:
        tokens = [t for t in texto.split()
                  if t not in _STOPWORDS_ES and len(t) > 2]
    return " ".join(tokens)


# ============================================================
# EXPORTAR NO CATEGORIZADOS
# ============================================================

def exportar_no_categorizados(
    df_original: pd.DataFrame,
    X_svd: np.ndarray,
    modelo,
    le,
    cfg: dict,
    log,
):
    """
    Evalúa TODOS los registros del dataset contra el modelo entrenado.
    Aplica dos criterios de baja confianza (combinados con OR):
      • LOW_PROB   : prob_max < umbral_prob_min
      • LOW_MARGIN : margen top1-top2 < umbral_margen_top
      • AMBOS      : los dos criterios fallan

    Exporta un Excel con dos pestañas:
      • Categorizados    → pasan ambos criterios
      • No_Categorizados → fallan uno o ambos → para revisión manual

    La pestaña No_Categorizados incluye columna 'Categoria_Correcta'
    (resaltada en amarillo) para que el revisor la complete y esas filas
    se incorporen al próximo ciclo de entrenamiento.
    """

    umbral_prob   = cfg["umbral_prob_min"]
    umbral_margen = cfg["umbral_margen_top"]
    ruta_salida   = Path(cfg["ruta_no_categ"])

    log.info("\n📋 EXPORTANDO NO CATEGORIZADOS")
    log.info("=" * 70)

    # ── Predicción con probabilidades sobre todo el dataset ──
    probs     = modelo.predict_proba(X_svd)           # (n, n_clases)
    top2_idx  = np.argsort(probs, axis=1)[:, -2:]     # top2, orden ascendente
    prob_top1 = probs[np.arange(len(probs)), top2_idx[:, -1]]
    prob_top2 = probs[np.arange(len(probs)), top2_idx[:, -2]]
    margen    = prob_top1 - prob_top2

    pred_clase = le.inverse_transform(top2_idx[:, -1])
    pred_top2  = le.inverse_transform(top2_idx[:, -2])

    # ── Criterios de baja confianza ───────────────────────────
    mask_low_prob   = prob_top1 < umbral_prob
    mask_low_margen = margen    < umbral_margen
    mask_no_categ   = mask_low_prob | mask_low_margen

    # ── Motivo legible ────────────────────────────────────────
    def _motivo(lp, lm):
        if lp and lm: return "AMBOS"
        if lp:        return "LOW_PROB"
        return               "LOW_MARGIN"

    motivos = [_motivo(lp, lm) for lp, lm in zip(mask_low_prob, mask_low_margen)]

    # ── Construir dataframe resultado ─────────────────────────
    df_result = df_original.copy().reset_index(drop=True)
    df_result["Categoria_Predicha"]  = pred_clase
    df_result["Categoria_Top2"]      = pred_top2
    df_result["Prob_Max"]            = np.round(prob_top1, 4)
    df_result["Prob_Top2"]           = np.round(prob_top2, 4)
    df_result["Margen_Top1_Top2"]    = np.round(margen,    4)
    df_result["Motivo_No_Categ"]     = motivos
    df_result["Baja_Confianza"]      = mask_no_categ
    df_result["Categoria_Correcta"]  = ""   # para revisión manual

    df_categ = (
        df_result[~mask_no_categ]
        .drop(columns=["Motivo_No_Categ", "Baja_Confianza"])
        .reset_index(drop=True)
    )
    df_no_categ = (
        df_result[mask_no_categ]
        .drop(columns=["Baja_Confianza"])
        .reset_index(drop=True)
    )

    n_total    = len(df_result)
    n_categ    = len(df_categ)
    n_no_categ = len(df_no_categ)
    pct        = n_no_categ / n_total * 100 if n_total > 0 else 0

    log.info(f"   • Umbral prob_max   : >= {umbral_prob}")
    log.info(f"   • Umbral margen     : >= {umbral_margen}")
    log.info(f"   • Total registros   : {n_total:,}")
    log.info(f"   • Categorizados     : {n_categ:,}  ({100-pct:.1f}%)")
    log.info(f"   • No categorizados  : {n_no_categ:,}  ({pct:.1f}%)")

    if n_no_categ > 0:
        log.info(f"\n   Distribución por motivo de rechazo:")
        serie_motivos = pd.Series(motivos)
        for mot, cnt in serie_motivos[mask_no_categ].value_counts().items():
            log.info(f"      • {mot:12}: {cnt:,} ({cnt/n_no_categ*100:.1f}%)")

        log.info(f"\n   Top categorías predichas con baja confianza:")
        for cat, cnt in df_no_categ["Categoria_Predicha"].value_counts().head(10).items():
            log.info(f"      • {str(cat):25}: {cnt:,}")

    # ── Exportar Excel con dos pestañas ───────────────────────
    ruta_salida.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(ruta_salida, engine="openpyxl") as writer:

        # Pestaña 1: Categorizados
        df_categ.to_excel(writer, sheet_name="Categorizados", index=False)
        ws_cat = writer.sheets["Categorizados"]
        _formatear_header(ws_cat)
        _autofit_columnas(ws_cat)

        # Pestaña 2: No_Categorizados
        if n_no_categ > 0:
            df_no_categ.to_excel(writer, sheet_name="No_Categorizados", index=False)
            ws_no = writer.sheets["No_Categorizados"]
            _formatear_header(ws_no)
            _resaltar_columna_correcta(ws_no, df_no_categ, "Categoria_Correcta")
            _autofit_columnas(ws_no)
        else:
            # Pestaña vacía con nota
            pd.DataFrame({"Nota": ["Sin registros de baja confianza en este ciclo"]}).to_excel(
                writer, sheet_name="No_Categorizados", index=False
            )

    kb = ruta_salida.stat().st_size / 1024
    log.info(f"\n   ✓ Excel exportado  : {ruta_salida}")
    log.info(f"   ✓ Tamaño           : {kb:.1f} KB")
    log.info(f"   ✓ Pestañas         : Categorizados | No_Categorizados")
    log.info(
        "\n   💡 Flujo de mejora continua:"
        "\n      1. Abre NO_CATEGORIZADOS.xlsx → pestaña No_Categorizados"
        "\n      2. Completa la columna amarilla 'Categoria_Correcta'"
        "\n      3. Pega esas filas en tu Data de Entrenamiento.xlsx"
        "\n         (columnas: Frase_clave + Categoria_Correcta como Categoría)"
        "\n      4. Vuelve a correr 1_ENTRENAR_MODELO.py"
        "\n      5. Compara F1-Macro nuevo vs anterior"
    )

    return df_categ, df_no_categ


# ── Helpers de formato openpyxl ──────────────────────────────

def _formatear_header(ws):
    header_fill = PatternFill("solid", fgColor="1A1D27")
    header_font = Font(color="00D4FF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

def _resaltar_columna_correcta(ws, df, nombre_col):
    if nombre_col not in df.columns:
        return
    col_idx   = df.columns.get_loc(nombre_col) + 1
    yellow    = PatternFill("solid", fgColor="FFF176")
    bold_font = Font(bold=True)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                            min_col=col_idx, max_col=col_idx):
        for cell in row:
            cell.fill = yellow
            cell.font = bold_font

def _autofit_columnas(ws):
    for col in ws.columns:
        max_len = max(
            (len(str(cell.value)) if cell.value is not None else 0)
            for cell in col
        )
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)


# ============================================================
# MAIN
# ============================================================

def main():
    log = setup_logger()

    log.info("=" * 70)
    log.info("🚀 ENTRENAMIENTO DE MODELO NLP TELCO")
    log.info(f"   {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    log.info("=" * 70)

    # ── 1. Cargar datos ───────────────────────────────────────
    ruta = Path(CFG["archivo_entrenamiento"])
    if not ruta.exists():
        log.error(f"❌ Archivo no encontrado: {ruta}")
        log.error("   Ajusta 'archivo_entrenamiento' en CFG")
        sys.exit(1)

    log.info(f"\n📂 Cargando: {ruta.name}")
    try:
        df = pd.read_excel(ruta, engine="openpyxl")
        log.info(f"   ✓ {len(df):,} filas | {len(df.columns)} columnas")
    except Exception as e:
        log.error(f"❌ Error leyendo Excel: {e}")
        sys.exit(1)

    for col in [CFG["col_texto"], CFG["col_categoria"]]:
        if col not in df.columns:
            log.error(f"❌ Columna '{col}' no encontrada")
            log.error(f"   Columnas disponibles: {df.columns.tolist()}")
            sys.exit(1)

    # ── 2. Limpiar datos ──────────────────────────────────────
    log.info("\n🧹 LIMPIANDO DATOS")
    log.info("=" * 70)

    df = df[[CFG["col_texto"], CFG["col_categoria"]]].copy()
    df = df.dropna(subset=[CFG["col_texto"], CFG["col_categoria"]])
    df[CFG["col_categoria"]] = df[CFG["col_categoria"]].str.strip().str.title()

    log.info(f"   • Registros válidos: {len(df):,}")
    log.info(f"\n   Distribución de categorías:")
    for cat, n in df[CFG["col_categoria"]].value_counts().items():
        log.info(f"      • {str(cat):25}: {n:6,} ({n/len(df)*100:5.1f}%)")

    conteo = df[CFG["col_categoria"]].value_counts()
    clases_validas = conteo[conteo >= CFG["min_muestras_clase"]].index
    descartadas    = conteo[conteo < CFG["min_muestras_clase"]]
    if len(descartadas) > 0:
        log.warning(f"\n   ⚠️  Categorías descartadas (< {CFG['min_muestras_clase']} muestras):")
        for cat, n in descartadas.items():
            log.warning(f"      • {cat}: {n} muestras")

    df = df[df[CFG["col_categoria"]].isin(clases_validas)].reset_index(drop=True)
    log.info(f"\n   ✓ Datos para entrenamiento: {len(df):,} | Categorías: {len(clases_validas)}")

    # ── 3. Preprocesar textos ─────────────────────────────────
    log.info("\n   Procesando textos...")
    if HAS_SPACY:
        log.info("   ✓ spaCy disponible — usando lematización")
    else:
        log.info("   ⚠️  spaCy no disponible — usando tokenización simple")

    df["texto_clean"] = df[CFG["col_texto"]].apply(
        lambda x: limpiar_texto(x, CFG["usar_lematizacion"])
    )
    df = df[df["texto_clean"].str.len() >= 5].reset_index(drop=True)
    log.info(f"   ✓ Textos procesados: {len(df):,}")

    # ── 4. TF-IDF ─────────────────────────────────────────────
    log.info("\n🔤 GENERANDO TF-IDF")
    log.info("=" * 70)

    vectorizer = TfidfVectorizer(
        max_features=CFG["tfidf_max_features"],
        ngram_range=CFG["tfidf_ngram_range"],
        min_df=CFG["tfidf_min_df"],
        max_df=CFG["tfidf_max_df"],
        sublinear_tf=True,
        strip_accents="unicode",
    )
    X_tfidf = vectorizer.fit_transform(df["texto_clean"])
    log.info(f"   ✓ Matriz TF-IDF: {X_tfidf.shape}")
    log.info(f"   ✓ Vocabulario  : {len(vectorizer.vocabulary_):,} términos")

    # ── 5. SVD ────────────────────────────────────────────────
    log.info("\n📉 REDUCCIÓN SVD")
    log.info("=" * 70)

    n_components = min(
        CFG["svd_components"],
        X_tfidf.shape[0] - 1,
        X_tfidf.shape[1] - 1
    )
    svd = TruncatedSVD(n_components=n_components, random_state=CFG["random_state"])
    X_svd = svd.fit_transform(X_tfidf)
    varianza = svd.explained_variance_ratio_.sum()
    log.info(f"   ✓ SVD: {X_svd.shape} | Varianza explicada: {varianza:.2%}")

    # ── 6. Codificar etiquetas ────────────────────────────────
    le = LabelEncoder()
    y  = le.fit_transform(df[CFG["col_categoria"]].values)
    log.info(f"\n   ✓ Categorías codificadas: {list(le.classes_)}")

    # ── 7. Split ──────────────────────────────────────────────
    X_train, X_test, y_train, y_test = train_test_split(
        X_svd, y,
        test_size=CFG["test_size"],
        random_state=CFG["random_state"],
        stratify=y
    )
    log.info(f"\n   ✓ Train: {len(y_train):,} | Test: {len(y_test):,}")

    # ── 8. SMOTE ──────────────────────────────────────────────
    if CFG["usar_smote"] and HAS_SMOTE:
        log.info("\n⚖️  APLICANDO SMOTE")
        log.info("=" * 70)
        try:
            min_clase = int(np.min(np.bincount(y_train)))
            k = max(1, min(5, min_clase - 1))
            smote = SMOTE(random_state=CFG["random_state"], k_neighbors=k)
            X_train, y_train = smote.fit_resample(X_train, y_train)
            log.info(f"   ✓ Muestras balanceadas: {len(y_train):,}")
        except Exception as e:
            log.warning(f"   ⚠️  SMOTE falló ({e}) — continuando sin balanceo")
    elif CFG["usar_smote"] and not HAS_SMOTE:
        log.warning("   ⚠️  SMOTE solicitado pero imbalanced-learn no instalado")
        log.warning("       Instala con: pip install imbalanced-learn")

    # ── 9. Entrenar modelo ────────────────────────────────────
    log.info("\n🤖 ENTRENANDO MODELO")
    log.info("=" * 70)

    modelo = LogisticRegression(
        max_iter=1000,
        random_state=CFG["random_state"],
        class_weight="balanced",
        solver="lbfgs",
    )
    modelo.fit(X_train, y_train)

    # ── 10. Evaluación ────────────────────────────────────────
    y_pred      = modelo.predict(X_test)
    acc         = accuracy_score(y_test, y_pred)
    f1_macro    = f1_score(y_test, y_pred, average="macro",    zero_division=0)
    f1_weighted = f1_score(y_test, y_pred, average="weighted", zero_division=0)

    log.info(f"   ✓ Accuracy    : {acc:.4f}")
    log.info(f"   ✓ F1-Macro    : {f1_macro:.4f}")
    log.info(f"   ✓ F1-Weighted : {f1_weighted:.4f}")

    cv = StratifiedKFold(n_splits=5, shuffle=True, random_state=CFG["random_state"])
    cv_scores = cross_val_score(modelo, X_train, y_train, cv=cv, scoring="f1_macro")
    log.info(f"   ✓ CV F1-Macro : {cv_scores.mean():.4f} ± {cv_scores.std():.4f}")

    log.info(f"\n   Reporte por categoría:")
    reporte = classification_report(
        y_test, y_pred,
        target_names=le.classes_,
        zero_division=0
    )
    for linea in reporte.split("\n"):
        if linea.strip():
            log.info(f"      {linea}")

    # ── 10b. Exportar no categorizados ────────────────────────
    # Opera sobre X_svd completo (todo el dataset, no solo test)
    # para que el Excel cubra todos los registros, no solo el 20%
    try:
        exportar_no_categorizados(
            df_original=df,
            X_svd=X_svd,
            modelo=modelo,
            le=le,
            cfg=CFG,
            log=log,
        )
    except Exception as e:
        log.warning(f"   ⚠️  Exportación no categorizados falló: {e}")
        log.warning(f"       {traceback.format_exc()}")

    # ── 11. Guardar PKL ───────────────────────────────────────
    log.info("\n💾 GUARDANDO MODELO")
    log.info("=" * 70)

    ruta_pkl = Path(CFG["ruta_salida_pkl"])
    ruta_pkl.parent.mkdir(parents=True, exist_ok=True)

    metadata = {
        "nombre":           "LogReg_TF-IDF_SVD",
        "accuracy":         round(acc, 4),
        "f1_macro":         round(f1_macro, 4),
        "f1_weighted":      round(f1_weighted, 4),
        "cv_f1_macro":      round(float(cv_scores.mean()), 4),
        "fecha":            datetime.now().isoformat(),
        "categorias":       le.classes_.tolist(),
        "n_entrenamiento":  len(df),
        "vocabulario":      len(vectorizer.vocabulary_),
        "svd_componentes":  n_components,
    }

    payload = {
        "vectorizer":    vectorizer,
        "svd":           svd,
        "modelo":        modelo,
        "label_encoder": le,
        "metadata":      metadata,
    }

    joblib.dump(payload, ruta_pkl)
    kb = ruta_pkl.stat().st_size / 1024
    log.info(f"   ✓ PKL guardado : {ruta_pkl}")
    log.info(f"   ✓ Tamaño       : {kb:.1f} KB")
    log.info(f"   ✓ Categorías   : {le.classes_.tolist()}")

    # ── 12. Gráficas ──────────────────────────────────────────
    log.info("\n📊 GENERANDO VISUALIZACIONES")
    log.info("=" * 70)

    ruta_graficas = Path(CFG["ruta_graficas"])
    ruta_graficas.mkdir(parents=True, exist_ok=True)

    for nombre_fn, fn, args in [
        ("Gráfica 1 - Distribución",     grafica_distribucion,      (df, CFG["col_categoria"], ruta_graficas, log)),
        ("Gráfica 2 - Confusión",        grafica_confusion,         (y_test, y_pred, le.classes_, ruta_graficas, log)),
        ("Gráfica 3 - F1 categoría",     grafica_f1_por_categoria,  (y_test, y_pred, le.classes_, ruta_graficas, log)),
        ("Gráfica 4 - Curva aprendizaje",grafica_curva_aprendizaje, (modelo, X_train, y_train, ruta_graficas, log)),
        ("Gráfica 5 - Dashboard",        grafica_dashboard,         (metadata, cv_scores, ruta_graficas, log)),
        ("Gráfica 6 - Top palabras",     grafica_top_palabras,      (vectorizer, modelo, le, ruta_graficas, log)),
    ]:
        try:
            fn(*args)
        except Exception as e:
            log.warning(f"   ⚠️  {nombre_fn} falló: {e}")

    log.info(f"\n   ✓ Gráficas guardadas en: {ruta_graficas}")

    log.info("\n" + "=" * 70)
    log.info("✅ ENTRENAMIENTO COMPLETADO")
    log.info(f"   PKL            → {ruta_pkl}")
    log.info(f"   No Categ Excel → {CFG['ruta_no_categ']}")
    log.info(f"   PNG            → {ruta_graficas}")
    log.info(f"   Siguiente paso → 2_categorizar_nps.py")
    log.info("=" * 70 + "\n")


if __name__ == "__main__":
    main()
